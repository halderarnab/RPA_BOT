from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class DataRow:
    row_number: int
    values: dict[str, Any]

    @property
    def row_id(self) -> str:
        invoice = str(self.values.get("invoice_number") or "").strip()
        return f"row-{self.row_number}-{invoice}" if invoice else f"row-{self.row_number}"


class ExcelDataReader:
    def __init__(self, aliases: dict[str, list[str]]) -> None:
        self.aliases = aliases

    def read_rows(self, workbook_path: Path) -> list[DataRow]:
        workbook = load_workbook(workbook_path, data_only=True)
        sheet = workbook.active
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = {
            self._normalize(cell.value): index
            for index, cell in enumerate(header_cells)
            if cell.value is not None
        }

        rows: list[DataRow] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            values: dict[str, Any] = {}
            for canonical, aliases in self.aliases.items():
                column_index = self._resolve_column(headers, aliases)
                if column_index is not None and column_index < len(row):
                    values[canonical] = self._format_cell(row[column_index])
            rows.append(DataRow(row_number=row_number, values=values))
        return rows

    def _resolve_column(self, headers: dict[str, int], aliases: list[str]) -> int | None:
        for alias in aliases:
            index = headers.get(self._normalize(alias))
            if index is not None:
                return index
        return None

    def _normalize(self, value: Any) -> str:
        return " ".join(str(value or "").strip().lower().split())

    def _format_cell(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")
        return value
